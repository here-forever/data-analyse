import csv
from collections.abc import Iterator
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from io import BytesIO, StringIO
from pathlib import Path

from openpyxl import load_workbook

from app.core.errors import AppError
from app.imports.schemas import FieldType, ImportFieldPreview

SUPPORTED_EXTENSIONS = {".csv", ".xlsx", ".xlsm"}
TabularSource = bytes | Path


class ParsedTabularFile:
    def __init__(
        self,
        *,
        file_type: str,
        fields: list[ImportFieldPreview],
        rows: list[dict[str, object | None]],
    ) -> None:
        self.file_type = file_type
        self.fields = fields
        self.rows = rows


def parse_tabular_file(file_name: str, content: bytes) -> ParsedTabularFile:
    extension = Path(file_name).suffix.lower()
    if extension not in SUPPORTED_EXTENSIONS:
        raise AppError(
            message="Only CSV and Excel files are supported",
            code="unsupported_file_type",
            status_code=400,
        )

    raw_rows = parse_csv(content) if extension == ".csv" else parse_excel(content)
    if not raw_rows:
        raise AppError(message="Uploaded file has no rows", code="empty_file", status_code=400)

    headers = normalize_headers(raw_rows[0])
    data_rows = [
        coerce_row(headers, row) for row in raw_rows[1:] if any(value is not None for value in row)
    ]
    fields = infer_fields(headers, data_rows)
    typed_rows = [coerce_typed_row(row, fields) for row in data_rows]

    return ParsedTabularFile(
        file_type=extension.removeprefix("."),
        fields=fields,
        rows=typed_rows,
    )


def iter_typed_tabular_rows(
    file_name: str,
    source: TabularSource,
    fields: list[ImportFieldPreview],
) -> Iterator[dict[str, object | None]]:
    """Stream typed rows from a retained source without building a full row list."""
    raw_rows = iter_raw_rows(file_name, source)
    try:
        headers = normalize_headers(next(raw_rows))
    except StopIteration as error:
        raise AppError(
            message="Uploaded file has no rows",
            code="empty_file",
            status_code=400,
        ) from error

    source_fields_by_order = {field.order: field for field in fields}
    for field in fields:
        if field.order >= len(headers):
            raise AppError(
                message="Dataset field order does not match source preview",
                code="invalid_dataset_field_order",
                status_code=400,
            )

    for raw_row in raw_rows:
        if not any(value is not None for value in raw_row):
            continue

        normalized_row = coerce_row(headers, raw_row)
        yield {
            source_fields_by_order[field.order].name: coerce_value(
                normalized_row.get(headers[field.order]),
                field.inferred_type,
            )
            for field in fields
        }


def iter_raw_rows(file_name: str, source: TabularSource) -> Iterator[list[object | None]]:
    extension = Path(file_name).suffix.lower()
    if extension not in SUPPORTED_EXTENSIONS:
        raise AppError(
            message="Only CSV and Excel files are supported",
            code="unsupported_file_type",
            status_code=400,
        )

    if extension == ".csv":
        if isinstance(source, bytes):
            text = source.decode("utf-8-sig")
            yield from (list(row) for row in csv.reader(StringIO(text)))
            return

        with source.open("r", encoding="utf-8-sig", newline="") as file_handle:
            yield from (list(row) for row in csv.reader(file_handle))
        return

    workbook_source = BytesIO(source) if isinstance(source, bytes) else source
    workbook = load_workbook(workbook_source, read_only=True, data_only=True)
    try:
        sheet = workbook.active
        yield from (list(row) for row in sheet.iter_rows(values_only=True))
    finally:
        workbook.close()


def parse_csv(content: bytes) -> list[list[object | None]]:
    text = content.decode("utf-8-sig")
    reader = csv.reader(StringIO(text))
    return [list(row) for row in reader]


def parse_excel(content: bytes) -> list[list[object | None]]:
    workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
    sheet = workbook.active
    return [list(row) for row in sheet.iter_rows(values_only=True)]


def normalize_headers(raw_headers: list[object | None]) -> list[str]:
    headers: list[str] = []
    seen: dict[str, int] = {}

    for index, raw_header in enumerate(raw_headers):
        fallback = f"column_{index + 1}"
        name = str(raw_header).strip() if raw_header is not None else fallback
        name = name or fallback
        count = seen.get(name, 0)
        seen[name] = count + 1
        headers.append(name if count == 0 else f"{name}_{count + 1}")

    return headers


def coerce_row(headers: list[str], raw_row: list[object | None]) -> dict[str, object | None]:
    return {
        header: normalize_cell(raw_row[index]) if index < len(raw_row) else None
        for index, header in enumerate(headers)
    }


def normalize_cell(value: object | None) -> object | None:
    if value is None:
        return None
    if isinstance(value, str):
        stripped = value.strip()
        return stripped or None
    return value


def infer_fields(
    headers: list[str],
    rows: list[dict[str, object | None]],
) -> list[ImportFieldPreview]:
    fields: list[ImportFieldPreview] = []
    for order, header in enumerate(headers):
        values = [row[header] for row in rows]
        non_empty_values = [value for value in values if value is not None]
        fields.append(
            ImportFieldPreview(
                name=header,
                inferred_type=infer_type(non_empty_values),
                nullable=len(non_empty_values) != len(values),
                order=order,
            )
        )
    return fields


def infer_type(values: list[object]) -> FieldType:
    if not values:
        return "text"
    if all(is_boolean(value) for value in values):
        return "boolean"
    if all(is_integer(value) for value in values):
        return "integer"
    if all(is_decimal(value) for value in values):
        return "decimal"
    if all(is_date(value) for value in values):
        return "date"
    if all(is_datetime(value) for value in values):
        return "datetime"
    return "text"


def is_boolean(value: object) -> bool:
    if isinstance(value, bool):
        return True
    return isinstance(value, str) and value.lower() in {"true", "false", "yes", "no", "1", "0"}


def is_integer(value: object) -> bool:
    if isinstance(value, bool):
        return False
    if isinstance(value, int):
        return True
    return isinstance(value, str) and value.isdecimal()


def is_decimal(value: object) -> bool:
    if isinstance(value, bool):
        return False
    if isinstance(value, int | float | Decimal):
        return True
    if not isinstance(value, str):
        return False
    try:
        Decimal(value)
    except InvalidOperation:
        return False
    return True


def is_date(value: object) -> bool:
    if isinstance(value, datetime):
        return False
    if isinstance(value, date):
        return True
    if not isinstance(value, str):
        return False
    try:
        datetime.strptime(value, "%Y-%m-%d")
    except ValueError:
        return False
    return True


def is_datetime(value: object) -> bool:
    if isinstance(value, datetime):
        return True
    if not isinstance(value, str):
        return False
    try:
        datetime.fromisoformat(value)
    except ValueError:
        return False
    return True


def coerce_typed_row(
    row: dict[str, object | None],
    fields: list[ImportFieldPreview],
) -> dict[str, object | None]:
    return {field.name: coerce_value(row[field.name], field.inferred_type) for field in fields}


def coerce_value(value: object | None, field_type: FieldType) -> object | None:
    if value is None:
        return None
    if field_type == "integer":
        return int(value)
    if field_type == "decimal":
        return float(value)
    if field_type == "boolean":
        return coerce_boolean(value)
    if field_type == "date":
        return value.isoformat() if isinstance(value, date) else str(value)
    if field_type == "datetime":
        return value.isoformat() if isinstance(value, datetime) else str(value)
    return str(value)


def coerce_boolean(value: object) -> bool:
    if isinstance(value, bool):
        return value
    return str(value).lower() in {"true", "yes", "1"}
